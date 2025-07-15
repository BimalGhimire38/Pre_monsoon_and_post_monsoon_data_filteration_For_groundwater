#2081/11/16  35:58 PM

import openpyxl
from openpyxl import Workbook,load_workbook

wb = load_workbook("01_STW_updated_with_recent_data.xlsx")
ws = wb.active

# r= ws.max_column
# print(r)
headers = [ws.cell(row=1,column=k).value for k in range (1,ws.max_column+1)]
all_data = {}
for i in range (2,ws.max_row+1):
    current_station = ws.cell(row=i,column=2).value
    if current_station not in all_data:
        all_data[current_station] = {}
    data_for_station = [ws.cell(row=i,column=k).value for k in range (1,ws.max_column+1)]
    for m in range(len(headers)):
        all_data[current_station][headers[m]] = data_for_station[m]

a = all_data.keys()
for i in a:
    print(i,"\n")

