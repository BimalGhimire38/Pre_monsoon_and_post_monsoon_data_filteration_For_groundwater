from openpyxl import load_workbook, Workbook

# Load the Excel file
file_path = "01_STW_updated_with_recent_data.xlsx"
wb = load_workbook(file_path)
ws = wb["All_STW"]  # Adjust the sheet name if different

# Get headers (assumed in first row)
headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

# Find columns that correspond to May and June
may_june_columns = [col for col in range(1, len(headers) + 1) if headers[col - 1] and (
    "-05-" in str(headers[col - 1]) or "-06-" in str(headers[col - 1])
)]

# Create a new workbook for filtered data
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title="May_June_Data"

# Copy headers (keeping necessary columns)
for idx, col in enumerate([1, 2, 3, 4] + may_june_columns, start=1):  # Keep first 3 columns (Station Name, X, Y)
    new_ws.cell(row=1, column=idx, value=ws.cell(row=1, column=col).value)

# Copy data for May and June
for row in range(2, ws.max_row + 1):
    for idx, col in enumerate([1, 2, 3] + may_june_columns, start=1):
        new_ws.cell(row=row, column=idx, value=ws.cell(row=row, column=col).value)

# Save the filtered data
filtered_file = "01_May_June_Data.xlsx"
new_wb.save(filtered_file)

print(f"Filtered data saved in {filtered_file}")
